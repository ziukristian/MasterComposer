import os
import time
from PIL import Image,ImageDraw,ImageFont
from openpyxl import load_workbook
import datetime
from pypdf import PdfMerger
import pandas as pd

ultimaRiga=1362 #PrimaRigaVuota TOGLI MARCO GIANESINI BLOCK
mode='n' #Cambia
anno=2023 #Cambia
luogo='Bussolengo'
meseParola='Novembre' #Cambia
mese=11 #Cambia
giorno=1

pathOutput="OUTPUT"
pathTemp="TEMP"
nomeFileStorico="StoricoPartecipanti.xlsx"
nomeFileListaPartecipanti="ListaPartecipanti.xlsx"
nomeFileCertificazioneVT="certificazioneVT.jpg"
nomeFileCertificazioneVS="certificazioneVS.jpg"
nomeFileCertificazioneN="certificazioneN.jpg"
nomeFileCertificazioneSA="certificazioneSA.jpg"
nomeFileCertificazioneMIV="certificazioneMIV.jpg"
nomeFontCavalieri="Bickham Script Pro Regular.ttf"
coloreFontCavalieri="#404040"
coloreTriangolo="#FF0000"
nomeFontCertificazioni="Bickham Script Pro Regular.ttf"
nomeFontNumeroCertificazioni="arial.ttf"
coloreFontCertificazioni="#000000"
coloreFontNumeroCertificazioni="#FFFFFF"
coloreFontCertificazioniMiv="#FFFFFF"
coloreFontNumeroCertificazioniMiv="#474747"
nomeFontAnnoCertificazioni="timesi.ttf"
nomeFontLuogoCertificazioni="times.ttf"
mivGenerato=False

fontAzienda = ImageFont.truetype("Montserrat-Regular.ttf", 150)
font = ImageFont.truetype("Montserrat-Regular.ttf", 180)
fontBold = ImageFont.truetype("Montserrat-Bold.ttf", 180)
fontSmallerBold = ImageFont.truetype("Montserrat-Bold.ttf", 150)
fontSmaller = ImageFont.truetype("Montserrat-Regular.ttf", 150)
coloreRifrequentante="#FF0000"
indexAzienda=13

def scrivi(testo,w,h,font,draw,hex="#000000"):
    W, H = (w, h)
    _, _, w, h = draw.textbbox((0, 0), testo, font=font)
    draw.text(((W - w) / 2, (H - h) / 2), testo, font=font, fill=hex)

def Esporta(partecipante,i):
    if partecipante.campus:
        im = Image.open("badge2.jpg")
    else:
        im = Image.open("badge.jpg")
    size = im.size
    draw = ImageDraw.Draw(im)


    if (partecipante.tipo == 'r'):
        draw.polygon([(size[0], 0), (size[0], size[0] * (27 / 100)), (size[0] - (size[0] * (27 / 100)), 0)],fill=coloreRifrequentante)


    azienda=partecipante.azienda.strip().upper()

    w, h = fontAzienda.getsize(azienda)

    if w > (size[0]*0.85):
        aziendaArray = azienda.split(' ',2)
    else:
        aziendaArray=[azienda]


    #nomepre=partecipante.nome.strip()+" "+partecipante.cognome.strip()

    #nome = nomepre.split(' ',1)

    #w, h = font.getsize(nome[1])
    #if w >(size[0]*0.85):
        #nome=nomepre.split(' ',2)


    #if (len(nome) == 1):
        #scrivi(nome[0], size[0], size[0], fontBold,draw)
    #if (len(nome) == 2):
        #scrivi(nome[0], size[0], size[0] * 1.03, fontBold,draw)
        #scrivi(nome[1], size[0], size[0] * 1.22, font,draw)
    #if (len(nome) == 3):
        #scrivi(nome[0], size[0], size[0] * 0.96, fontBold,draw)
        #scrivi(nome[1], size[0], size[0] * 1.12, font,draw)
        #scrivi(nome[2], size[0], size[0] * 1.28, font,draw)

    nome=partecipante.nome.strip().upper()
    cognome=partecipante.cognome.strip().upper()
    w, h = font.getsize(nome)
    if w >(size[0]*0.85):
        scrivi(nome, size[0], size[0] * 1.03, fontSmallerBold,draw)
    else:
        scrivi(nome, size[0], size[0] * 1.03, fontBold, draw)

    w, h = font.getsize(partecipante.cognome.strip())
    if w > (size[0] * 0.85):
        scrivi(cognome, size[0], size[0] * 1.22, fontSmaller, draw)
    else:
        scrivi(cognome, size[0], size[0] * 1.22, font, draw)

    if (len(aziendaArray) == 1):
        scrivi(aziendaArray[0], size[0], size[0] * 1.76, fontAzienda,draw)
    if (len(aziendaArray) == 2):
        scrivi(aziendaArray[0], size[0], size[0] * 1.63, fontAzienda,draw)
        scrivi(aziendaArray[1], size[0], size[0] * 1.76, fontAzienda,draw)
    if (len(aziendaArray) == 3):
        fontDaUsare=fontAzienda
        w, h = fontAzienda.getsize(aziendaArray[0])
        if w > (size[0] * 0.85):
            fontDaUsare=ImageFont.truetype("Montserrat-Regular.ttf", 90)
        scrivi(aziendaArray[0], size[0], size[0] * 1.63, fontDaUsare,draw)
        w, h = fontAzienda.getsize(aziendaArray[1])
        if w > (size[0] * 0.85):
            fontDaUsare = ImageFont.truetype("Montserrat-Regular.ttf", 90)
        scrivi(aziendaArray[1], size[0], size[0] * 1.76, fontDaUsare,draw)
        w, h = fontAzienda.getsize(aziendaArray[2])
        if w > (size[0] * 0.85):
            fontDaUsare = ImageFont.truetype("Montserrat-Regular.ttf", 90)
        scrivi(aziendaArray[2], size[0], size[0] * 1.89, fontDaUsare,draw)

    im.save('result'+str(i)+'.pdf')
    im.close()
    merger.append('result' + str(i) + '.pdf')

def trovaFontSize(img,fontname,testo):
    draw = ImageDraw.Draw(img)
    size=500
    font = ImageFont.truetype(fontname, size)
    box = draw.textbbox((0, 0), testo, font=font)
    text_width = box[2]
    while text_width > (img.size[0]*0.80):
        size=size-5
        font = ImageFont.truetype(fontname, size)
        box = draw.textbbox((0, 0), testo, font=font)
        text_width = box[2]
    return size

def generaFile(lista):
    global mivGenerato
    if not os.path.exists(pathTemp):
        os.makedirs(pathTemp)
    lista.sort(key=lambda x: x.nome)
    mergerCavalieri = PdfMerger()
    mergerCavalieriR = PdfMerger()
    mergerModulo = PdfMerger()
    mergerMiv = PdfMerger()

    for i, par in enumerate(lista):
        test=True
        #if par.nome + ' ' + par.cognome == "Marco Gianesini":
            #test = False
        print(par.nome)
        if mode != 'sa':
            #GENERO CAVALIERE
            font = ImageFont.truetype("Bickham Script Pro Regular.ttf", 310)
            if par.campus:
                imgCavaliere = Image.open('cavaliere2.jpg')
            else:
                imgCavaliere = Image.open('cavaliere.jpg')

            drawCavaliere = ImageDraw.Draw(imgCavaliere)
            testo = (par.nome if par.nome!= None else '') + ' ' + (par.cognome if par.cognome!= None else '')
            boxCognome = drawCavaliere.textbbox((0, 0), par.cognome, font=font)
            boxNome = drawCavaliere.textbbox((0, 0), par.nome, font=font)
            widthCognome=boxCognome[2]
            widthNome = boxNome[2]


            if widthNome > imgCavaliere.width * 0.8 or widthCognome > imgCavaliere.width * 0.8:
                font = ImageFont.truetype("Bickham Script Pro Regular.ttf", 295)
            box = drawCavaliere.textbbox((0, 0), testo, font=font)
            textWidth = box[2]
            textHeight = box[3]


            #linea orizzontale
            drawCavaliere.line([(0,imgCavaliere.height/2),(imgCavaliere.width,imgCavaliere.height/2)], fill="#dbdbdb")

            #Se supera 80% dell immagine vado a capo
            if textWidth > imgCavaliere.width * 0.8:
                #Inserisco il nome
                #Capovolto
                testo = par.nome if par.nome != None else ''
                box = drawCavaliere.textbbox((0, 0), testo, font=font)
                nome_textWidth = box[2]
                nome_textHeight = box[3]
                capovolto = Image.new('RGBA', (round(nome_textWidth*1.2), nome_textHeight), (0, 0, 0, 0))
                drawCapovolto = ImageDraw.Draw(capovolto)
                drawCapovolto.text((round((nome_textWidth*1.2-nome_textWidth)/2), 0), testo, font=font, fill=("#404040"))
                capovolto = capovolto.rotate(180)
                x = round((imgCavaliere.width - round(nome_textWidth*1.2)) / 2)
                y = round(imgCavaliere.height * 0.46 - nome_textHeight)
                imgCavaliere.paste(capovolto, (x, y), capovolto)
                #Dritto
                x = round((imgCavaliere.width - nome_textWidth) / 2)
                y = round((imgCavaliere.height * 0.53))
                drawCavaliere.text((x, y), testo, font=font, fill=("#404040"))

                # Inserisco il cognome capovolto
                # Capovolto
                testo = par.cognome if par.cognome != None else ''
                box = drawCavaliere.textbbox((0, 0), testo, font=font)
                cognome_textWidth = box[2]
                cognome_textHeight = box[3]
                capovolto = Image.new('RGBA', (round(cognome_textWidth*1.2), cognome_textHeight), (0, 0, 0, 0))
                drawCapovolto = ImageDraw.Draw(capovolto)
                drawCapovolto.text((round((cognome_textWidth*1.2-cognome_textWidth)/2), 0), testo, font=font, fill=("#404040"))
                capovolto = capovolto.rotate(180)
                x = round((imgCavaliere.width - round(cognome_textWidth*1.2)) / 2)
                y = round(imgCavaliere.height * 0.37 - cognome_textHeight)
                imgCavaliere.paste(capovolto, (x, y), capovolto)
                # Dritto
                x = round((imgCavaliere.width - cognome_textWidth) / 2)
                y = round((imgCavaliere.height * 0.62))
                drawCavaliere.text((x, y), testo, font=font, fill=("#404040"))
            else:

                # Capovolto
                capovolto = Image.new('RGBA', (round(textWidth*1.2), textHeight), (0, 0, 0, 0))
                drawCapovolto = ImageDraw.Draw(capovolto)
                drawCapovolto.text((round((textWidth*1.2-textWidth)/2), 0), testo, font=font, fill=("#404040"))
                capovolto = capovolto.rotate(180, expand=1)
                x = round((imgCavaliere.width - round(textWidth*1.2)) / 2)
                y = round(imgCavaliere.height * 0.37 - textHeight)
                imgCavaliere.paste(capovolto, (x, y), capovolto)
                # Dritto
                x = round((imgCavaliere.width - textWidth) / 2)
                y = round((imgCavaliere.height * 0.61))
                drawCavaliere.text((x, y), testo, font=font, fill=("#404040"))
            #Genero triangolo per rifrequentanti
            if par.tipo == 'r':
                print("ROSSO CAVALIERE ---------------------------")
                drawCavaliere.polygon(
                    [(0, imgCavaliere.height / 2), (imgCavaliere.width * 0.20, imgCavaliere.height / 2), (0, (imgCavaliere.height / 2) - (imgCavaliere.width * 0.20))],
                    fill=("#cb242a"))
                drawCavaliere.polygon([(imgCavaliere.width, imgCavaliere.height / 2), (imgCavaliere.width * 0.80, imgCavaliere.height / 2),
                            (imgCavaliere.width, (imgCavaliere.height / 2) + (imgCavaliere.width * 0.20))], fill=("#cb242a"))

            imgCavaliere.save(pathTemp + '/cav' + str(i) + '.pdf')
            imgCavaliere.close()
            if par.tipo == 'r':
                mergerCavalieriR.append(pathTemp + '/cav' + str(i) + '.pdf')
            else:
                mergerCavalieri.append(pathTemp + '/cav' + str(i) + '.pdf')

        #GENERO CERTIFICAZIONE MODULO
        if mode == 'vt':
            imgModulo = Image.open("certificazioneVT.jpg")
            numero = str(par.nVT)
        if mode == 'vs':
            imgModulo = Image.open("certificazioneVS.jpg")
            numero = str(par.nVS)
        if mode == 'n':
            imgModulo = Image.open("certificazioneN.jpg")
            numero = str(par.nN)
        if mode == 'sa':
            imgModulo = Image.open("certificazioneSA.jpg")
            numero = str(par.nN)

        drawModulo = ImageDraw.Draw(imgModulo)
        testo = (par.nome if par.nome != None else '') + ' ' + (par.cognome if par.cognome != None else '')
        size=trovaFontSize(imgModulo, "Bickham Script Pro Regular.ttf", testo)
        font = ImageFont.truetype("Bickham Script Pro Regular.ttf",size )
        box = drawModulo.textbbox((0, 0), testo, font=font)
        textWidth = box[2]
        textHeight = box[3]
        x = round((imgModulo.width - textWidth) / 2)
        y = round((imgModulo.height * 0.2165)+((500-size)/1.6))
        drawModulo.text((x, y), testo, font=font, fill=("#000000"))
        if par.tipo == 'r' and int(numero)>1:
            font = ImageFont.truetype("arial.ttf", 69)
            box = drawModulo.textbbox((0, 0), numero, font=font)
            textWidth = box[2]
            textHeight = box[3]
            x = round((imgModulo.width - textWidth) / 2)
            y = round(imgModulo.height * 0.627)
            drawModulo.text((x, y), numero, font=font, fill=("#FFFFFF"))

        #Scrivo luogo e data
        font = ImageFont.truetype("times.ttf", 69)
        drawModulo.text((imgModulo.width * 0.7, imgModulo.height * 0.88), luogo + ' - ', font=font, fill=(coloreFontCertificazioni))
        box = drawModulo.textbbox((0, 0), luogo + ' - ', font=font)
        textWidth = box[2]
        textHeight = box[3]
        font = ImageFont.truetype("timesi.ttf", 69)
        drawModulo.text((imgModulo.width * 0.7 + textWidth, imgModulo.height * 0.88), meseParola + ' ' + str(anno), font=font,fill=("#000000"))

        #Scrivo Anno
        font = ImageFont.truetype("arial.ttf", 120)
        box = drawModulo.textbbox((0, 0), str(anno), font=font)
        textWidth = box[2]
        textHeight = box[3]
        drawModulo.text((((imgModulo.width - textWidth) / 2), imgModulo.height * 0.905), str(anno), font=font,fill=(coloreFontNumeroCertificazioni))

        imgModulo.save(pathTemp + '/cert' + str(i) + '.pdf')
        imgModulo.close()
        if(test):
            mergerModulo.append(pathTemp + '/cert' + str(i) + '.pdf')

        #GENERO MIV se lo fa
        if par.nuovoMIV and mode != 'sa' and test:

            mivGenerato=True
            verticale=0.958
            imgMiv = Image.open("certificazioneMIV.jpg")
            drawMiv = ImageDraw.Draw(imgMiv)
            testo = (par.nome if par.nome != None else '') + ' ' + (par.cognome if par.cognome != None else '')
            font = ImageFont.truetype("Bickham Script Pro Regular.ttf", trovaFontSize(imgMiv, "Bickham Script Pro Regular.ttf", testo))
            box = drawModulo.textbbox((0, 0), testo, font=font)
            textWidth = box[2]
            textHeight = box[3]
            x = round((imgMiv.width - textWidth) / 2)
            y = round(imgMiv.height * 0.21)
            drawMiv.text((x, y), testo, font=font, fill=("#FFFFFF"))
            font = ImageFont.truetype("arial.ttf", 80)

            if(par.nMIV>1):
                testo = str(par.nMIV)

                box = drawModulo.textbbox((0, 0), testo, font=font)
                textWidth = box[2]
                textHeight = box[3]
                x = round((imgMiv.width - textWidth) / 2)
                y = round(imgMiv.height * 0.37)
                drawMiv.text((x, y), testo, font=font, fill=("#474747"))

            box = drawModulo.textbbox((0, 0), luogo, font=font)
            textWidth = box[2]
            textHeight = box[3]
            drawMiv.text((round((imgMiv.width - textWidth) / 2), imgMiv.height * verticale), luogo, font=font,fill=("#FFFFFF"))

            box = drawModulo.textbbox((0, 0), str(anno), font=font)
            textWidth = box[2]
            drawMiv.text((round((imgMiv.width - textWidth) / 2), imgMiv.height * verticale + textHeight), str(anno), font=font,fill=("#FFFFFF"))

            imgMiv.save(pathTemp + '/miv' + str(i) + '.pdf')
            imgMiv.close()

            mergerMiv.append(pathTemp + '/miv' + str(i) + '.pdf')

        #GENERO BADGE
        #Esporta(par, i)

    mergerCavalieri.write(pathOutput + "/" + time.strftime("%Y%m%d") + "_Cavalieri.pdf")
    mergerCavalieriR.write(pathOutput + "/" + time.strftime("%Y%m%d") + "_CavalieriRifrequentanti.pdf")
    mergerModulo.write(pathOutput + "/" + time.strftime("%Y%m%d") + "_CertificazioniModulo.pdf")
    if mivGenerato:
        mergerMiv.write(pathOutput + "/" + time.strftime("%Y%m%d") + "_CertificazioniMIV.pdf")

class Partecipante:
  def __init__(self, nome, cognome,mail,telefono,tipo,mode,campus,azienda):
    self.nome = nome
    self.cognome = cognome
    self.mail = mail
    self.telefono = telefono
    self.tipo = tipo
    self.campus = campus
    self.azienda =  '' if pd.isnull(azienda) else azienda
    self.nVT = 1 if mode=='vt' else 0
    self.nVS = 1 if mode=='vs' else 0
    self.nN = 1 if mode=='n' else 0
    self.nMIV = 0
    self.nuovoMIV = False
    self.daInserire = True

  def __repr__(self):
    return (
        f"NOME: {self.nome}\n"
        f"COGNOME: {self.cognome}\n"
        f"nVT: {self.nVT}\n"
        f"nVS: {self.nVS}\n"
        f"nN: {self.nN}\n"
        f"nMIV: {self.nMIV}\n"
        f"nuovoMIV: {self.nuovoMIV}\n\n"
        f"campus: {self.campus}\n\n"
    )


wbStorico = load_workbook(filename = nomeFileStorico,data_only=True)
wsStorico = wbStorico.active
wbPartecipanti = load_workbook(filename = nomeFileListaPartecipanti,data_only=True)
wsPartecipanti = wbPartecipanti.active

if not os.path.exists(pathOutput):
    os.makedirs(pathOutput)

def trovaPartecipanti(ws):
    # Indice meno uno
    partecipanti=[]
    index = 53
    if mode=='vt':
        index = 51
    if mode == 'vs':
        index = 52
    if mode == 'sa':
        index = 52
    for row in ws.iter_rows(values_only=True):
        campus=False
        if row[116]==1:
            campus=True
        if mode !='sa' and row[112]=='SI' and row[113]!='SI':
            if row[index]=='R':
                newEle=Partecipante(row[6],row[7],row[8],row[10],'r',mode,campus,row[indexAzienda])
                partecipanti.append(newEle)
            if row[index] == 'F':
                newEle=Partecipante(row[6],row[7],row[8],row[10],'f',mode,campus,row[indexAzienda])
                partecipanti.append(newEle)
        if mode =='sa' and (row[index]==1 or row[index]=='P') and row[index+1]==1:
            newEle = Partecipante(row[1], row[2], row[3], row[4], 'f', mode,campus,row[indexAzienda])
            partecipanti.append(newEle)
    return partecipanti

def aggiungiRiga(par):
    global ultimaRiga
    wsStorico.cell(row=ultimaRiga, column=3).value=par.nome
    wsStorico.cell(row=ultimaRiga, column=4).value = par.cognome
    wsStorico.cell(row=ultimaRiga, column=5).value = (par.nome if par.nome != None else '')+' '+(par.cognome if par.cognome != None else '')
    if mode=='vt':
        wsStorico.cell(row=ultimaRiga, column=6).value = datetime.datetime(anno,mese,giorno)
    if mode=='vs':
        wsStorico.cell(row=ultimaRiga, column=7).value = datetime.datetime(anno,mese,giorno)
    if mode=='n':
        wsStorico.cell(row=ultimaRiga, column=8).value = datetime.datetime(anno,mese,giorno)
    wsStorico.cell(row=ultimaRiga, column=11).value = par.mail
    wsStorico.cell(row=ultimaRiga, column=10).value = 'Rifrequenza' if par.tipo=='r' else ''
    ultimaRiga= ultimaRiga+1

def contaPartecipazioni(lista,wsStorico):
    for par in lista:
        #if par.nome + ' ' + par.cognome == "Marco Gianesini":
            #print("SALTATOOOOOOOOOOOOOOOOOOOOOOOO")
            #continue
        if (par.nome != None):
            par.nome = par.nome.rstrip()
        if (par.cognome != None):
            par.nome = par.nome.rstrip()
        i=0
        for row in wsStorico.iter_rows(values_only=True):
            if par.nome==row[2] and par.cognome==row[3]:

                if par.mail != row[10]:
                    #Scrivi su file nome cognome per anomalia
                    f = open(pathOutput+"/"+time.strftime("%Y%m%d")+"_MailDiverse.txt", "w")
                    f.write('INDICE: '+str(i)+'\nNOME: '+(par.nome if par.nome != None else '')+'\nCOGNOME: '+(par.cognome if par.cognome != None else '')+
                            '\nNUOVA MAIL: '+(par.mail if par.mail != None else '')+'\nVECCHIA MAIL: '+(row[10] if row[10] != None else '')+'\n\n')

                if row[5] != None:
                    par.nVT=par.nVT+1
                elif par.daInserire and mode=='vt':
                    if row[6] != None and row[7] != None :
                        par.nuovoMIV = True
                        par.nMIV=par.nMIV+1
                        d=wsStorico.cell(row=i + 1, column=9)
                        d.number_format = '0'
                        d.value = 1
                    wsStorico.cell(row=i+1, column=6).value=datetime.datetime(anno,mese,giorno)
                    par.daInserire=False
                if row[6] != None:
                    par.nVS=par.nVS+1
                elif par.daInserire and mode=='vs':
                    if row[5] != None and row[7] != None :
                        par.nuovoMIV = True
                        par.nMIV = par.nMIV + 1
                        d=wsStorico.cell(row=i + 1, column=9)
                        d.number_format = '0'
                        d.value = 1
                    wsStorico.cell(row=i+1, column=7).value=datetime.datetime(anno,mese,giorno)
                    par.daInserire=False
                if row[7] != None:
                    par.nN=par.nN+1
                elif par.daInserire and mode=='n':
                    if row[5] != None and row[6] != None :
                        par.nuovoMIV = True
                        par.nMIV = par.nMIV + 1
                        d=wsStorico.cell(row=i + 1, column=9)
                        d.number_format = '0'
                        d.value = 1
                    wsStorico.cell(row=i+1, column=8).value=datetime.datetime(anno,mese,giorno)
                    par.daInserire=False
                if row[8] != None:
                    par.nMIV=par.nMIV+1
            i = i + 1
        if par.daInserire:
            aggiungiRiga(par)

#partecipanti=trovaPartecipanti(wsPartecipanti)
#partecipanti.sort(key=lambda x: x.nome.strip()+x.cognome.strip())
#if mode != 'sa':
    #contaPartecipazioni(partecipanti,wsStorico)
    #wbStorico.save(pathOutput+"/"+time.strftime("%Y%m%d")+"_StoricoPartecipanti.xlsx")
#f = open(pathOutput+"/Partecipanti.txt", "w")
#for x in partecipanti:
    #f.write(f"NOME: {x.nome}\n"
        #f"COGNOME: {x.cognome}\n"
        #f"Num Tattica: {x.nVT}\n"
        #f"Num Strategica: {x.nVS}\n"
        #f"Num Negoziazione: {x.nN}\n"
        #f"Num MIV: {x.nMIV}\n"
        #f"Conclude MIV: {x.nuovoMIV}\n"
        #f"campus: {x.campus}\n\n")

merger = PdfMerger()
#generaFile(partecipanti)


partec=[Partecipante("Silvia","Sardena","null","",'r',mode,1,""),Partecipante("Edoardo","Botteghi","null","",'f',mode,1,""),Partecipante("Andrea","Simonato","null","",'f',mode,1,"")]



silvia = Partecipante("Silvia","Sardena","null","",'r',mode,0,"")
silvia.nN = 2
simo = Partecipante("Andrea","Simonato","null","",'f',mode,0,"")
simo.nuovoMIV = True


generaFile([silvia,Partecipante("Edoardo","Botteghi","null","",'f',mode,0,""),simo])

badge=Partecipante("Diego","Guerrieri","null","",'f',mode,0,"SERFORM")
Esporta(badge,0)
badge=Partecipante("Silvia","Sardena","null","",'f',mode,0,"")
Esporta(badge,1)
badge=Partecipante("Edoardo","Botteghi","null","",'f',mode,0,"Spiagge")
Esporta(badge,2)
badge=Partecipante("Andrea","Simonato","null","",'f',mode,0,"Steelform")
Esporta(badge,3)
merger.write("badgeCollection.pdf")
